#!/usr/bin/env python3
"""Build Agni Health unit economics Excel workbook."""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import os

OUTPUT = "/Users/sohanai/.openclaw/workspace/agni-health/planning/unit-economics-panel.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Unit Economics"

# ── Palette ───────────────────────────────────────────────────────────────────
YELLOW_FILL   = PatternFill("solid", fgColor="FFFACD")   # input cells
GRAY_FILL     = PatternFill("solid", fgColor="F2F2F2")   # computed cells
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # section headers
SUBHDR_FILL   = PatternFill("solid", fgColor="2E5090")   # panel sub-headers
COL_HDR_FILL  = PatternFill("solid", fgColor="4472C4")   # table column headers
GREEN_FILL    = PatternFill("solid", fgColor="C6EFCE")   # $/hr ≥ 500

WHITE_FONT    = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT     = Font(bold=True, size=10)
NORM_FONT     = Font(size=10)
COMPUTED_FONT = Font(italic=True, size=10, color="444444")

thin = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_fill(fill):
    return fill

def style_header(cell, fill=HEADER_FILL):
    cell.fill = fill
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

def style_input(cell):
    cell.fill = YELLOW_FILL
    cell.font = NORM_FONT
    cell.alignment = Alignment(horizontal="right")
    cell.border = THIN_BORDER

def style_computed(cell):
    cell.fill = GRAY_FILL
    cell.font = COMPUTED_FONT
    cell.alignment = Alignment(horizontal="right")
    cell.border = THIN_BORDER

def style_label(cell, bold=False):
    cell.font = BOLD_FONT if bold else NORM_FONT
    cell.alignment = Alignment(horizontal="left", indent=1)
    cell.border = THIN_BORDER

def style_table_cell(cell, is_computed=True):
    cell.fill = GRAY_FILL if is_computed else YELLOW_FILL
    cell.font = NORM_FONT
    cell.alignment = Alignment(horizontal="right")
    cell.border = THIN_BORDER

# ── Column widths ─────────────────────────────────────────────────────────────
#   A: row labels, B: Essential, C: Intensive
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16

# ── ROW TRACKER ───────────────────────────────────────────────────────────────
r = 1

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 HEADER
# ══════════════════════════════════════════════════════════════════════════════
ws.merge_cells(f"A{r}:C{r}")
ws.row_dimensions[r].height = 22
c = ws[f"A{r}"]
c.value = "SECTION 1 — ASSUMPTIONS (edit yellow cells)"
style_header(c)
r += 1

# Column sub-headers
ws[f"A{r}"].value = "Assumption"
ws[f"B{r}"].value = "Essential"
ws[f"C{r}"].value = "Intensive"
for col in ("A", "B", "C"):
    style_header(ws[f"{col}{r}"], fill=COL_HDR_FILL)
ws.row_dimensions[r].height = 18
r += 1

# Helper to add an assumption row
def add_assumption(label, val_e, val_i, computed=False, bold=False, fmt="#,##0.0"):
    global r
    ws[f"A{r}"].value = label
    style_label(ws[f"A{r}"], bold=bold)
    if computed:
        ws[f"B{r}"].value = val_e   # formula string
        ws[f"C{r}"].value = val_i
        style_computed(ws[f"B{r}"])
        style_computed(ws[f"C{r}"])
        ws[f"B{r}"].number_format = fmt
        ws[f"C{r}"].number_format = fmt
    else:
        ws[f"B{r}"].value = val_e
        ws[f"C{r}"].value = val_i
        style_input(ws[f"B{r}"])
        style_input(ws[f"C{r}"])
        if "$" in label or "Price" in label:
            ws[f"B{r}"].number_format = '"$"#,##0'
            ws[f"C{r}"].number_format = '"$"#,##0'
        else:
            ws[f"B{r}"].number_format = "#,##0"
            ws[f"C{r}"].number_format = "#,##0"
    rr = r
    r += 1
    return rr

# ── Assumptions ───────────────────────────────────────────────────────────────
r_price      = add_assumption("Price ($/mo per patient)",               300,  600)
r_init_len   = add_assumption("Initial visit length (min)",              60,   60)
r_init_amo   = add_assumption("Initial visit amortized over (months)",   12,   12)

# → Amortized initial (computed)
r_amoInit = r
ws[f"A{r}"].value = "→ Amortized initial visit (min/mo)"
style_label(ws[f"A{r}"])
ws[f"B{r}"].value = f"=B{r_init_len}/B{r_init_amo}"
ws[f"C{r}"].value = f"=C{r_init_len}/C{r_init_amo}"
style_computed(ws[f"B{r}"]); style_computed(ws[f"C{r}"])
ws[f"B{r}"].number_format = "#,##0.0"
ws[f"C{r}"].number_format = "#,##0.0"
r += 1

r_fu_len   = add_assumption("Follow-up visit length (min)",              20,   20)
r_fu_freq  = add_assumption("Follow-up visits per year",                  4,   12)

# → Follow-up time (min/mo)
r_fuTime = r
ws[f"A{r}"].value = "→ Follow-up time (min/mo)"
style_label(ws[f"A{r}"])
ws[f"B{r}"].value = f"=(B{r_fu_len}*B{r_fu_freq})/12"
ws[f"C{r}"].value = f"=(C{r_fu_len}*C{r_fu_freq})/12"
style_computed(ws[f"B{r}"]); style_computed(ws[f"C{r}"])
ws[f"B{r}"].number_format = "#,##0.0"
ws[f"C{r}"].number_format = "#,##0.0"
r += 1

r_lab    = add_assumption("Lab review per patient (min/mo)",              5,    8)
r_async  = add_assumption("Async messaging (min/mo)",                     5,   10)
r_admin  = add_assumption("Admin / coordination (min/mo)",                3,    5)

# → TOTAL time (min/mo)
r_totMin = r
ws[f"A{r}"].value = "→ TOTAL time per patient (min/mo)"
style_label(ws[f"A{r}"], bold=True)
ws[f"B{r}"].value = f"=B{r_amoInit}+B{r_fuTime}+B{r_lab}+B{r_async}+B{r_admin}"
ws[f"C{r}"].value = f"=C{r_amoInit}+C{r_fuTime}+C{r_lab}+C{r_async}+C{r_admin}"
style_computed(ws[f"B{r}"]); style_computed(ws[f"C{r}"])
ws[f"B{r}"].number_format = "#,##0.0"
ws[f"C{r}"].number_format = "#,##0.0"
ws[f"B{r}"].font = Font(bold=True, italic=True, size=10, color="444444")
ws[f"C{r}"].font = Font(bold=True, italic=True, size=10, color="444444")
r += 1

# → TOTAL time (hrs/mo)
r_totHr = r
ws[f"A{r}"].value = "→ TOTAL time per patient (hrs/mo)"
style_label(ws[f"A{r}"], bold=True)
ws[f"B{r}"].value = f"=B{r_totMin}/60"
ws[f"C{r}"].value = f"=C{r_totMin}/60"
style_computed(ws[f"B{r}"]); style_computed(ws[f"C{r}"])
ws[f"B{r}"].number_format = "#,##0.00"
ws[f"C{r}"].number_format = "#,##0.00"
ws[f"B{r}"].font = Font(bold=True, italic=True, size=10, color="444444")
ws[f"C{r}"].font = Font(bold=True, italic=True, size=10, color="444444")
r += 1

# → Revenue per hour
r_rph = r
ws[f"A{r}"].value = "→ Revenue per hour ($/hr)"
style_label(ws[f"A{r}"], bold=True)
ws[f"B{r}"].value = f"=B{r_price}/B{r_totHr}"
ws[f"C{r}"].value = f"=C{r_price}/C{r_totHr}"
style_computed(ws[f"B{r}"]); style_computed(ws[f"C{r}"])
ws[f"B{r}"].number_format = '"$"#,##0.00'
ws[f"C{r}"].number_format = '"$"#,##0.00'
ws[f"B{r}"].font = Font(bold=True, italic=True, size=10, color="1F3864")
ws[f"C{r}"].font = Font(bold=True, italic=True, size=10, color="1F3864")
r += 1

# Spacer
r += 1

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 HEADER
# ══════════════════════════════════════════════════════════════════════════════
# Expand columns A–G for the table
for col_letter, width in zip(["A","B","C","D","E","F","G"],
                              [20, 14, 14, 12, 16, 16, 14]):
    ws.column_dimensions[col_letter].width = max(
        ws.column_dimensions[col_letter].width, width
    )

ws.merge_cells(f"A{r}:G{r}")
ws.row_dimensions[r].height = 22
c = ws[f"A{r}"]
c.value = "SECTION 2 — PANEL MIX SCENARIOS (all formula-driven)"
style_header(c)
r += 1

# Table column headers
tbl_headers = ["Mix", "Essential pts", "Intensive pts", "Total pts", "Hrs/mo", "MRR ($)", "$/hr"]
for idx, h in enumerate(tbl_headers):
    col = get_column_letter(idx + 1)
    cell = ws[f"{col}{r}"]
    cell.value = h
    style_header(cell, fill=COL_HDR_FILL)
ws.row_dimensions[r].height = 18
r += 1

# ── Panel scenario builder ────────────────────────────────────────────────────
mixes = [
    ("100% Essential",  1.0, 0.0),
    ("80 / 20",         0.8, 0.2),
    ("60 / 40",         0.6, 0.4),
    ("50 / 50",         0.5, 0.5),
    ("40 / 60",         0.4, 0.6),
    ("20 / 80",         0.2, 0.8),
    ("100% Intensive",  0.0, 1.0),
]

dollar_sign_rule_ranges = []

for panel_size in [50, 60]:
    # Sub-header for panel size
    ws.merge_cells(f"A{r}:G{r}")
    ws.row_dimensions[r].height = 18
    c = ws[f"A{r}"]
    c.value = f"Panel Size: {panel_size} patients"
    c.fill = SUBHDR_FILL
    c.font = WHITE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = THIN_BORDER
    r += 1

    for mix_label, frac_e, frac_i in mixes:
        pts_e = round(panel_size * frac_e)
        pts_i = round(panel_size * frac_i)

        # Col A: label
        ws[f"A{r}"].value = mix_label
        style_label(ws[f"A{r}"])

        # Col B: Essential pts (hard-coded numbers, not formulas, but reference panel)
        ws[f"B{r}"].value = pts_e
        style_table_cell(ws[f"B{r}"], is_computed=False)
        ws[f"B{r}"].number_format = "#,##0"

        # Col C: Intensive pts
        ws[f"C{r}"].value = pts_i
        style_table_cell(ws[f"C{r}"], is_computed=False)
        ws[f"C{r}"].number_format = "#,##0"

        # Col D: Total pts = B+C
        ws[f"D{r}"].value = f"=B{r}+C{r}"
        style_table_cell(ws[f"D{r}"])
        ws[f"D{r}"].number_format = "#,##0"

        # Col E: Hrs/mo = E_pts * hrs_E + I_pts * hrs_I
        ws[f"E{r}"].value = f"=B{r}*B${r_totHr}+C{r}*C${r_totHr}"
        style_table_cell(ws[f"E{r}"])
        ws[f"E{r}"].number_format = "#,##0.0"

        # Col F: MRR = E_pts*price_E + I_pts*price_I
        ws[f"F{r}"].value = f"=B{r}*B${r_price}+C{r}*C${r_price}"
        style_table_cell(ws[f"F{r}"])
        ws[f"F{r}"].number_format = '"$"#,##0'

        # Col G: $/hr = MRR / Hrs_mo
        ws[f"G{r}"].value = f"=IF(E{r}>0,F{r}/E{r},\"\")"
        style_table_cell(ws[f"G{r}"])
        ws[f"G{r}"].number_format = '"$"#,##0.00'

        dollar_sign_rule_ranges.append(f"G{r}")
        r += 1

    r += 1  # spacer between panels

# ── Conditional formatting: $/hr >= 500 → green ───────────────────────────────
from openpyxl.formatting.rule import CellIsRule
green_rule = CellIsRule(
    operator="greaterThanOrEqual",
    formula=["500"],
    fill=GREEN_FILL,
    font=Font(color="276221", bold=True)
)
for cell_addr in dollar_sign_rule_ranges:
    ws.conditional_formatting.add(cell_addr, green_rule)

# ── Freeze panes just below section headers ────────────────────────────────────
ws.freeze_panes = "A3"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
