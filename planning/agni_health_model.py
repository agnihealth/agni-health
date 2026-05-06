#!/usr/bin/env python3
"""
Agni Health Financial Model - 500 Patients in 18 Months
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Constants
MONTHLY_PRICE = 199
ANNUAL_PRICE_MONTHLY = 174  # $2,088/yr = $174/mo
BLENDED_MRR = 191.50
MONTHLY_MIX = 0.70
ANNUAL_MIX = 0.30
PAYMENT_PROCESSING = 0.029
CAC = 200
ANNUAL_CHURN = 0.08
MONTHLY_CHURN = 0.0067  # 8% annual ≈ 0.67%/month
REFERRAL_RATE = 0.20  # 20% referrals after month 6
CONSULT_FEE = 275
TARGET_PATIENTS = 500
TARGET_MONTHS = 18

# Operating costs (annual, converted to monthly where needed)
MALPRACTICE_ANNUAL = 10000
EHR_ANNUAL = 2000
LICENSING_ANNUAL = 8000
ACCOUNTING_LEGAL_ANNUAL = 8000
SOFTWARE_TELEHEALTH_ANNUAL = 5000
BUSINESS_INSURANCE_MISC_ANNUAL = 5000
TOTAL_FIXED_ANNUAL = (MALPRACTICE_ANNUAL + EHR_ANNUAL + LICENSING_ANNUAL + 
                       ACCOUNTING_LEGAL_ANNUAL + SOFTWARE_TELEHEALTH_ANNUAL + 
                       BUSINESS_INSURANCE_MISC_ANNUAL)
MONTHLY_FIXED_COSTS = TOTAL_FIXED_ANNUAL / 12

MA_SALARY_ANNUAL = 50000
MA_HIRE_THRESHOLD = 150

def calculate_patient_ramp():
    """
    Back-calculate the patient acquisition ramp needed to hit 500 in 18 months.
    We need to account for churn and solve for required new patients each month.
    """
    # Strategy: Ramp up aggressively, then taper as referrals kick in
    # Target curve: start slow (month 1-2), ramp hard (month 3-12), maintain (13-18)
    
    # First, let's simulate different growth scenarios to find one that works
    # We'll use an S-curve approach with aggressive middle growth
    
    patients = 0
    monthly_data = []
    
    # Define target new patients per month (will adjust)
    # Need roughly 500 / 18 = ~28 avg, but with churn we need more
    # Churn will lose about 0.67% per month
    
    # Start with a growth curve and iterate
    base_targets = [
        15, 20, 28, 32, 35, 38,   # Months 1-6: ramp up
        40, 42, 38, 35, 32, 30,   # Months 7-12: peak then taper (referrals help)
        28, 25, 22, 20, 18, 15    # Months 13-18: coast with referrals
    ]
    
    # Simulate and adjust
    for attempt in range(20):
        patients = 0
        monthly_data = []
        
        for month in range(1, 19):
            # Calculate churn from existing patients
            churned = round(patients * MONTHLY_CHURN, 1)
            
            # Target new patients for this month
            target_new = base_targets[month - 1]
            
            # After month 6, 20% are referrals (no CAC)
            if month > 6:
                referral_new = round(target_new * REFERRAL_RATE)
                paid_new = target_new - referral_new
            else:
                referral_new = 0
                paid_new = target_new
            
            # Update patient count
            patients = patients - churned + target_new
            
            monthly_data.append({
                'month': month,
                'new_paid': paid_new,
                'new_referral': referral_new,
                'new_total': target_new,
                'churned': churned,
                'cumulative': patients
            })
        
        final_patients = patients
        
        # Adjust targets if we're not hitting 500
        if final_patients < 495:
            # Scale up
            scale = 500 / final_patients
            base_targets = [int(t * scale * 1.02) for t in base_targets]
        elif final_patients > 510:
            # Scale down
            scale = 500 / final_patients
            base_targets = [int(t * scale * 0.98) for t in base_targets]
        else:
            break
    
    return monthly_data, base_targets

def build_financial_model():
    """Build the complete 18-month financial model."""
    
    patient_data, new_patient_targets = calculate_patient_ramp()
    
    model = []
    cumulative_patients = 0
    cumulative_pnl = 0
    
    for month in range(1, 19):
        data = patient_data[month - 1]
        
        # Patient metrics
        new_paid = data['new_paid']
        new_referral = data['new_referral']
        new_total = data['new_total']
        churned = data['churned']
        cumulative_patients = data['cumulative']
        
        # Revenue
        consult_revenue = new_total * CONSULT_FEE
        membership_revenue = cumulative_patients * BLENDED_MRR
        gross_revenue = consult_revenue + membership_revenue
        
        # Costs
        processing_fees = gross_revenue * PAYMENT_PROCESSING
        cac_spend = new_paid * CAC
        
        # Marketing spend (we'll estimate based on CAC multiple for awareness)
        # Assume marketing = 1.5x CAC spend for brand building early, tapering later
        if month <= 6:
            marketing_multiplier = 1.5
        elif month <= 12:
            marketing_multiplier = 1.2
        else:
            marketing_multiplier = 0.8
        
        marketing_spend = cac_spend * marketing_multiplier
        
        # Fixed operating costs
        fixed_costs = MONTHLY_FIXED_COSTS
        
        # MA cost (after 150 patients)
        if cumulative_patients >= MA_HIRE_THRESHOLD:
            ma_cost = MA_SALARY_ANNUAL / 12
        else:
            ma_cost = 0
        
        # Total costs
        total_costs = processing_fees + cac_spend + marketing_spend + fixed_costs + ma_cost
        
        # Net income
        net_income = gross_revenue - total_costs
        cumulative_pnl += net_income
        
        model.append({
            'month': month,
            'new_paid': new_paid,
            'new_referral': new_referral,
            'new_total': new_total,
            'churned': round(churned, 1),
            'cumulative_patients': round(cumulative_patients, 0),
            'marketing_spend': round(marketing_spend, 2),
            'cac_spend': round(cac_spend, 2),
            'consult_revenue': round(consult_revenue, 2),
            'membership_revenue': round(membership_revenue, 2),
            'gross_revenue': round(gross_revenue, 2),
            'processing_fees': round(processing_fees, 2),
            'fixed_costs': round(fixed_costs, 2),
            'ma_cost': round(ma_cost, 2),
            'total_costs': round(total_costs, 2),
            'net_income': round(net_income, 2),
            'cumulative_pnl': round(cumulative_pnl, 2)
        })
    
    return model

def create_excel_file(model):
    """Create the Excel workbook with all tabs."""
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    money_format = '$#,##0'
    money_format_cents = '$#,##0.00'
    percent_format = '0.0%'
    number_format = '#,##0'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternating row colors
    light_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # =====================
    # TAB 1: Month-by-Month Model
    # =====================
    ws1 = wb.active
    ws1.title = "Monthly Model"
    
    # Headers
    headers = [
        'Month', 'New Patients (Paid)', 'New Patients (Referral)', 'Total New',
        'Churned', 'Cumulative Patients', 'Marketing Spend', 'CAC Spend',
        'Consult Revenue', 'Membership Revenue', 'Gross Revenue',
        'Processing Fees', 'Fixed Costs', 'MA Cost', 'Total Costs',
        'Net Income', 'Cumulative P&L'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    for row_idx, data in enumerate(model, 2):
        row_data = [
            data['month'],
            data['new_paid'],
            data['new_referral'],
            data['new_total'],
            data['churned'],
            int(data['cumulative_patients']),
            data['marketing_spend'],
            data['cac_spend'],
            data['consult_revenue'],
            data['membership_revenue'],
            data['gross_revenue'],
            data['processing_fees'],
            data['fixed_costs'],
            data['ma_cost'],
            data['total_costs'],
            data['net_income'],
            data['cumulative_pnl']
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            
            # Apply number formats
            if col in [7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]:  # Money columns
                cell.number_format = money_format
            
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = light_fill
            
            # Color net income
            if col == 16:
                if value >= 0:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
    
    # Totals row
    total_row = len(model) + 2
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    
    total_cols = {
        2: sum(d['new_paid'] for d in model),
        3: sum(d['new_referral'] for d in model),
        4: sum(d['new_total'] for d in model),
        5: sum(d['churned'] for d in model),
        7: sum(d['marketing_spend'] for d in model),
        8: sum(d['cac_spend'] for d in model),
        9: sum(d['consult_revenue'] for d in model),
        10: sum(d['membership_revenue'] for d in model),
        11: sum(d['gross_revenue'] for d in model),
        12: sum(d['processing_fees'] for d in model),
        13: sum(d['fixed_costs'] for d in model),
        14: sum(d['ma_cost'] for d in model),
        15: sum(d['total_costs'] for d in model),
        16: sum(d['net_income'] for d in model),
    }
    
    for col, value in total_cols.items():
        cell = ws1.cell(row=total_row, column=col, value=round(value, 2))
        cell.font = Font(bold=True)
        cell.border = thin_border
        if col >= 7:
            cell.number_format = money_format
    
    # Column widths
    col_widths = [8, 15, 18, 10, 10, 18, 15, 12, 15, 18, 15, 15, 12, 10, 12, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze panes
    ws1.freeze_panes = 'B2'
    
    # =====================
    # TAB 2: Summary
    # =====================
    ws2 = wb.create_sheet("Summary")
    
    # Find key milestones
    breakeven_month = None
    ma_hire_month = None
    self_sustaining_month = None
    
    for data in model:
        if data['net_income'] >= 0 and breakeven_month is None:
            breakeven_month = data['month']
        if data['cumulative_patients'] >= MA_HIRE_THRESHOLD and ma_hire_month is None:
            ma_hire_month = data['month']
        if data['cumulative_pnl'] >= 0 and self_sustaining_month is None:
            self_sustaining_month = data['month']
    
    # Year 1 (months 1-12) and Year 2 (months 13-18, partial) metrics
    year1_revenue = sum(d['gross_revenue'] for d in model[:12])
    year1_costs = sum(d['total_costs'] for d in model[:12])
    year1_net = year1_revenue - year1_costs
    
    year2_revenue = sum(d['gross_revenue'] for d in model[12:])
    year2_costs = sum(d['total_costs'] for d in model[12:])
    year2_net = year2_revenue - year2_costs
    
    total_marketing = sum(d['marketing_spend'] for d in model)
    total_cac = sum(d['cac_spend'] for d in model)
    final_mrr = model[-1]['cumulative_patients'] * BLENDED_MRR
    final_arr = final_mrr * 12
    
    # Header
    ws2.cell(row=1, column=1, value="AGNI HEALTH - FINANCIAL SUMMARY").font = Font(bold=True, size=16)
    ws2.cell(row=2, column=1, value="Target: 500 Patients in 18 Months").font = Font(italic=True, size=12)
    
    # Key Assumptions section
    ws2.cell(row=4, column=1, value="KEY ASSUMPTIONS").font = Font(bold=True, size=12)
    ws2.cell(row=4, column=1).fill = header_fill
    ws2.cell(row=4, column=1).font = Font(bold=True, color="FFFFFF", size=12)
    
    assumptions = [
        ("Monthly membership price", f"${MONTHLY_PRICE}/mo (70% of patients)"),
        ("Annual membership price", f"${ANNUAL_PRICE_MONTHLY}/mo (30% of patients)"),
        ("Blended MRR per patient", f"${BLENDED_MRR}"),
        ("Initial consultation fee", f"${CONSULT_FEE}"),
        ("Customer Acquisition Cost (CAC)", f"${CAC}"),
        ("Annual churn rate", f"{ANNUAL_CHURN*100}% ({MONTHLY_CHURN*100:.2f}%/month)"),
        ("Referral rate (after month 6)", f"{REFERRAL_RATE*100:.0f}% of new patients"),
        ("Payment processing", f"{PAYMENT_PROCESSING*100}%"),
        ("MA hire threshold", f"{MA_HIRE_THRESHOLD} patients"),
    ]
    
    for i, (label, value) in enumerate(assumptions, 5):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=value)
    
    # Key Milestones section
    row = 15
    ws2.cell(row=row, column=1, value="KEY MILESTONES").font = Font(bold=True, size=12)
    ws2.cell(row=row, column=1).fill = header_fill
    ws2.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=12)
    
    milestones = [
        ("Break-even month (monthly profit ≥ $0)", f"Month {breakeven_month}" if breakeven_month else "Not reached"),
        ("MA hire month (≥150 patients)", f"Month {ma_hire_month}" if ma_hire_month else "Not reached"),
        ("Self-sustaining (cumulative P&L ≥ $0)", f"Month {self_sustaining_month}" if self_sustaining_month else "Not reached"),
        ("Final patient count (Month 18)", f"{int(model[-1]['cumulative_patients'])} patients"),
    ]
    
    for i, (label, value) in enumerate(milestones, row + 1):
        ws2.cell(row=i, column=1, value=label)
        cell = ws2.cell(row=i, column=2, value=value)
        cell.font = Font(bold=True)
    
    # Financial Summary section
    row = 21
    ws2.cell(row=row, column=1, value="FINANCIAL SUMMARY").font = Font(bold=True, size=12)
    ws2.cell(row=row, column=1).fill = header_fill
    ws2.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF", size=12)
    
    financials = [
        ("Total Marketing Spend (18 months)", total_marketing, money_format),
        ("Total CAC Spend (18 months)", total_cac, money_format),
        ("Total Customer Acquisition (Marketing + CAC)", total_marketing + total_cac, money_format),
        ("", "", None),
        ("Year 1 Gross Revenue (Months 1-12)", year1_revenue, money_format),
        ("Year 1 Total Costs", year1_costs, money_format),
        ("Year 1 Net Income", year1_net, money_format),
        ("", "", None),
        ("Year 2 Partial Revenue (Months 13-18)", year2_revenue, money_format),
        ("Year 2 Partial Costs", year2_costs, money_format),
        ("Year 2 Partial Net Income", year2_net, money_format),
        ("", "", None),
        ("18-Month Total Revenue", year1_revenue + year2_revenue, money_format),
        ("18-Month Total Costs", year1_costs + year2_costs, money_format),
        ("18-Month Net Income", year1_net + year2_net, money_format),
        ("", "", None),
        ("Monthly Recurring Revenue at 500 patients", final_mrr, money_format),
        ("Annualized Revenue Run-Rate at 500 patients", final_arr, money_format),
    ]
    
    for i, (label, value, fmt) in enumerate(financials, row + 1):
        ws2.cell(row=i, column=1, value=label)
        cell = ws2.cell(row=i, column=2, value=value if value != "" else "")
        if fmt:
            cell.number_format = fmt
        if "Net Income" in label:
            cell.font = Font(bold=True)
            if isinstance(value, (int, float)):
                cell.fill = green_fill if value >= 0 else red_fill
    
    # Column widths
    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 35
    
    # =====================
    # TAB 3: Marketing Ramp
    # =====================
    ws3 = wb.create_sheet("Marketing Ramp")
    
    # Headers
    headers = ['Month', 'Target New Patients', 'Paid Acquisitions', 'Referrals',
               'CAC Spend', 'Marketing Spend', 'Total Acquisition Cost', 
               'Cumulative Patients', 'Cost per Total Patient']
    
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    cumulative_acq_cost = 0
    for row_idx, data in enumerate(model, 2):
        cumulative_acq_cost += data['marketing_spend'] + data['cac_spend']
        cost_per_patient = cumulative_acq_cost / data['cumulative_patients'] if data['cumulative_patients'] > 0 else 0
        
        row_data = [
            data['month'],
            data['new_total'],
            data['new_paid'],
            data['new_referral'],
            data['cac_spend'],
            data['marketing_spend'],
            data['marketing_spend'] + data['cac_spend'],
            int(data['cumulative_patients']),
            cost_per_patient
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            
            if col in [5, 6, 7, 9]:
                cell.number_format = money_format
            
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    # Totals row
    total_row = len(model) + 2
    ws3.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    
    total_new = sum(d['new_total'] for d in model)
    total_paid = sum(d['new_paid'] for d in model)
    total_referrals = sum(d['new_referral'] for d in model)
    total_cac_spend = sum(d['cac_spend'] for d in model)
    total_marketing_spend = sum(d['marketing_spend'] for d in model)
    total_acq = total_cac_spend + total_marketing_spend
    
    totals = [None, total_new, total_paid, total_referrals, total_cac_spend, 
              total_marketing_spend, total_acq, None, total_acq / model[-1]['cumulative_patients']]
    
    for col, value in enumerate(totals, 1):
        if value is not None:
            cell = ws3.cell(row=total_row, column=col, value=round(value, 2) if isinstance(value, float) else value)
            cell.font = Font(bold=True)
            cell.border = thin_border
            if col in [5, 6, 7, 9]:
                cell.number_format = money_format
    
    # Add marketing strategy notes
    note_row = total_row + 3
    ws3.cell(row=note_row, column=1, value="MARKETING STRATEGY NOTES").font = Font(bold=True, size=12)
    ws3.cell(row=note_row, column=1).fill = header_fill
    ws3.cell(row=note_row, column=1).font = Font(bold=True, color="FFFFFF", size=12)
    
    notes = [
        "• Months 1-6: Heavy marketing investment (1.5x CAC) for brand building and awareness",
        "• Months 7-12: Moderate marketing (1.2x CAC) as referrals begin contributing 20% of new patients",
        "• Months 13-18: Reduced marketing (0.8x CAC) as referrals and word-of-mouth drive growth",
        f"• Total marketing + CAC investment: ${total_acq:,.0f}",
        f"• Effective cost per acquired patient: ${total_acq / model[-1]['cumulative_patients']:,.0f}",
        f"• Referral savings: {total_referrals} patients acquired at $0 CAC",
    ]
    
    for i, note in enumerate(notes, note_row + 1):
        ws3.cell(row=i, column=1, value=note)
    
    # Column widths
    col_widths = [8, 18, 16, 10, 12, 15, 18, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze panes
    ws3.freeze_panes = 'B2'
    
    # Save workbook
    filepath = "/Users/sohanai/.openclaw/workspace/agni-health-500-18mo.xlsx"
    wb.save(filepath)
    print(f"Excel file saved to: {filepath}")
    
    # Print summary
    print("\n" + "="*60)
    print("AGNI HEALTH FINANCIAL MODEL SUMMARY")
    print("="*60)
    print(f"\nTarget: {TARGET_PATIENTS} patients in {TARGET_MONTHS} months")
    print(f"Final patient count: {int(model[-1]['cumulative_patients'])}")
    print(f"\nKey Milestones:")
    print(f"  • Break-even month: {breakeven_month}")
    print(f"  • MA hire month: {ma_hire_month}")
    print(f"  • Self-sustaining: {self_sustaining_month}")
    print(f"\nFinancial Summary:")
    print(f"  • Total Marketing + CAC: ${total_acq:,.0f}")
    print(f"  • Year 1 Net Income: ${year1_net:,.0f}")
    print(f"  • Year 2 (6 mo) Net Income: ${year2_net:,.0f}")
    print(f"  • 18-Month Net Income: ${year1_net + year2_net:,.0f}")
    print(f"  • MRR at 500 patients: ${final_mrr:,.0f}")
    print(f"  • ARR at 500 patients: ${final_arr:,.0f}")

if __name__ == "__main__":
    model = build_financial_model()
    create_excel_file(model)
