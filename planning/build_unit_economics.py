from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Unit Economics"

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SEC_FILL   = PatternFill("solid", fgColor="2F5496")   # section blue
SUB_FILL   = PatternFill("solid", fgColor="D6E4F7")   # light blue
CALC_FILL  = PatternFill("solid", fgColor="EAF2FB")   # very light blue
RESULT_FILL= PatternFill("solid", fgColor="E2EFDA")   # light green for results

HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SEC_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(name="Calibri", size=10)
BOLD_FONT  = Font(name="Calibri", bold=True, size=10)
RESULT_FONT= Font(name="Calibri", bold=True, size=10)

thin = Side(style="thin", color="B8CCE4")
med  = Side(style="medium", color="2F5496")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BOTTOM  = Border(left=thin, right=thin, top=thin, bottom=med)

def cell(ws, row, col, value=None, font=None, fill=None, align="left",
         border=None, num_fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    if font:  c.font = font
    elif bold: c.font = Font(name="Calibri", bold=True, size=10)
    if fill:  c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if border: c.border = border
    if num_fmt: c.number_format = num_fmt
    return c

def section_header(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SEC_FONT
    c.fill = SEC_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = MED_BOTTOM
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 18

def row_label(ws, row, label, ess_val, int_val,
              fill=None, bold=False, num_fmt='0.0', note=None):
    f = fill or PatternFill()
    fn = BOLD_FONT if bold else LABEL_FONT
    cell(ws, row, 1, label,    font=fn,   fill=f, border=THIN_BORDER)
    cell(ws, row, 2, ess_val,  font=fn,   fill=f, border=THIN_BORDER, align="right", num_fmt=num_fmt)
    cell(ws, row, 3, int_val,  font=fn,   fill=f, border=THIN_BORDER, align="right", num_fmt=num_fmt)
    cell(ws, row, 4, note or "", font=Font(name="Calibri", size=9, italic=True, color="595959"),
         fill=f, border=THIN_BORDER)

# ── Column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 36

# ── Title row ─────────────────────────────────────────────────────────────────
r = 1
ws.row_dimensions[r].height = 26
c = ws.cell(row=r, column=1, value="Agni Health — Physician Unit Economics")
c.font = Font(name="Calibri", bold=True, size=14, color="1F3864")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

r = 2
ws.row_dimensions[r].height = 6

# ── Column headers ────────────────────────────────────────────────────────────
r = 3
ws.row_dimensions[r].height = 20
for col, txt in [(1,"Activity / Metric"), (2,"Essential ($300/mo)"), (3,"Intensive ($600/mo)"), (4,"Notes")]:
    c = ws.cell(row=r, column=col, value=txt)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center" if col>1 else "left", vertical="center")
    c.border = THIN_BORDER
ws.row_dimensions[r].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Time per patient/month (minutes)
# ══════════════════════════════════════════════════════════════════════════════
r = 4; section_header(ws, r, "1.  Time per Patient / Month  (minutes)")

r = 5
row_label(ws, r, "Initial visit (amortized ÷ 12)", 5.0, 5.0,
          note="60 min visit / 12 months")

r = 6
row_label(ws, r, "Follow-up visits", 6.7, 20.0,
          note="Essential: quarterly 20 min (÷3 ≈ 6.7/mo); Intensive: monthly 20 min")

r = 7
row_label(ws, r, "Lab review & ordering", 5.0, 8.0,
          note="Essential: ~quarterly; Intensive: ~monthly")

r = 8
row_label(ws, r, "Async messaging (48hr / 24hr SLA)", 5.0, 10.0,
          note="Essential: lighter volume; Intensive: heavier")

r = 9
row_label(ws, r, "Care coordination / admin", 3.0, 5.0,
          note="Referrals, prior auths, care team coordination")

# Total row — formula-driven
r = 10
ws.row_dimensions[r].height = 18
for col in range(1, 5):
    ws.cell(row=r, column=col).fill = SUB_FILL
    ws.cell(row=r, column=col).border = MED_BOTTOM

ws.cell(row=r, column=1, value="TOTAL minutes / patient / month").font = BOLD_FONT
ws.cell(row=r, column=1).fill = SUB_FILL
ws.cell(row=r, column=1).border = MED_BOTTOM
ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")

for col, rng in [(2, "B5:B9"), (3, "C5:C9")]:
    c = ws.cell(row=r, column=col, value=f"=SUM({rng})")
    c.font = BOLD_FONT
    c.fill = SUB_FILL
    c.border = MED_BOTTOM
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = "0.0"

ws.cell(row=r, column=4).fill = SUB_FILL
ws.cell(row=r, column=4).border = MED_BOTTOM

# Store total-row reference for later formulas
TOT_ESS = "B10"
TOT_INT = "C10"

r = 11; ws.row_dimensions[r].height = 6

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — Revenue per hour
# ══════════════════════════════════════════════════════════════════════════════
r = 12; section_header(ws, r, "2.  Revenue per Hour  (single patient)")

r = 13
row_label(ws, r, "Monthly price per patient", 300, 600,
          fill=CALC_FILL, num_fmt='"$"#,##0', note="Sohan sets this")

r = 14
# hours/patient/month = total_min / 60
for col, ref in [(2, TOT_ESS), (3, TOT_INT)]:
    c = ws.cell(row=r, column=col, value=f"={ref}/60")
    c.font = LABEL_FONT
    c.fill = CALC_FILL
    c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = "0.00"
ws.cell(row=r, column=1, value="Hours / patient / month").font = LABEL_FONT
ws.cell(row=r, column=1).fill = CALC_FILL
ws.cell(row=r, column=1).border = THIN_BORDER
ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws.cell(row=r, column=4, value="total min ÷ 60").font = Font(name="Calibri", size=9, italic=True, color="595959")
ws.cell(row=r, column=4).fill = CALC_FILL
ws.cell(row=r, column=4).border = THIN_BORDER

r = 15
ws.row_dimensions[r].height = 18
for col in range(1, 5):
    ws.cell(row=r, column=col).fill = RESULT_FILL
    ws.cell(row=r, column=col).border = MED_BOTTOM

ws.cell(row=r, column=1, value="Revenue / hour (single patient rate)").font = RESULT_FONT
ws.cell(row=r, column=1).fill = RESULT_FILL
ws.cell(row=r, column=1).border = MED_BOTTOM
ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")

for col, price_ref, hr_ref in [(2,"B13","B14"), (3,"C13","C14")]:
    c = ws.cell(row=r, column=col, value=f"={price_ref}/{hr_ref}")
    c.font = RESULT_FONT
    c.fill = RESULT_FILL
    c.border = MED_BOTTOM
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = '"$"#,##0'

ws.cell(row=r, column=4, value="price ÷ hrs per patient").font = Font(name="Calibri", size=9, italic=True, color="595959")
ws.cell(row=r, column=4).fill = RESULT_FILL
ws.cell(row=r, column=4).border = MED_BOTTOM

r = 16; ws.row_dimensions[r].height = 6

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — Panel economics
# ══════════════════════════════════════════════════════════════════════════════
r = 17; section_header(ws, r, "3.  Panel Economics")

# Sub-header
r = 18
ws.row_dimensions[r].height = 16
for col, txt in [(1,"Scenario"), (2,"Total hrs/mo"), (3,"Gross MRR"), (4,"$/hr earned")]:
    c = ws.cell(row=r, column=col, value=txt)
    c.font = Font(name="Calibri", bold=True, size=9, color="1F3864")
    c.fill = PatternFill("solid", fgColor="BDD7EE")
    c.alignment = Alignment(horizontal="center" if col>1 else "left", vertical="center")
    c.border = THIN_BORDER

# Helper: build a panel row
# mins_per_patient_ess = B10, mins_per_patient_int = C10
# price_ess=300, price_int=600

def panel_row(ws, row, label, n_ess, n_int, fill=None):
    f = fill or PatternFill()
    ws.row_dimensions[row].height = 16
    ws.cell(row=row, column=1, value=label).font = LABEL_FONT
    ws.cell(row=row, column=1).fill = f
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # total hrs/mo = (n_ess * B10 + n_int * C10) / 60
    hrs_formula = f"=({n_ess}*{TOT_ESS}+{n_int}*{TOT_INT})/60"
    c = ws.cell(row=row, column=2, value=hrs_formula)
    c.font = LABEL_FONT; c.fill = f; c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = "0.0"

    # gross MRR = n_ess*300 + n_int*600
    mrr_formula = f"={n_ess}*300+{n_int}*600"
    c = ws.cell(row=row, column=3, value=mrr_formula)
    c.font = LABEL_FONT; c.fill = f; c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = '"$"#,##0'

    # $/hr = MRR / hrs
    dph_formula = f"=C{row}/B{row}"
    c = ws.cell(row=row, column=4, value=dph_formula)
    c.font = RESULT_FONT; c.fill = RESULT_FILL; c.border = THIN_BORDER
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.number_format = '"$"#,##0'

# ── 50-patient panel ──
r = 19
ws.row_dimensions[r].height = 14
c = ws.cell(row=r, column=1, value="── 50 patients ──")
c.font = Font(name="Calibri", bold=True, size=9, italic=True, color="2F5496")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

r = 20; panel_row(ws, r, "50 pts · 100% Essential  (50 Ess, 0 Int)",  50, 0)
r = 21; panel_row(ws, r, "50 pts · 50/50 mix        (25 Ess, 25 Int)", 25, 25, fill=CALC_FILL)
r = 22; panel_row(ws, r, "50 pts · 100% Intensive   (0 Ess, 50 Int)",   0, 50)

# ── 60-patient panel ──
r = 23
ws.row_dimensions[r].height = 14
c = ws.cell(row=r, column=1, value="── 60 patients ──")
c.font = Font(name="Calibri", bold=True, size=9, italic=True, color="2F5496")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

r = 24; panel_row(ws, r, "60 pts · 100% Essential  (60 Ess, 0 Int)",  60, 0)
r = 25; panel_row(ws, r, "60 pts · 50/50 mix        (30 Ess, 30 Int)", 30, 30, fill=CALC_FILL)
r = 26; panel_row(ws, r, "60 pts · 100% Intensive   (0 Ess, 60 Int)",   0, 60)

r = 27; ws.row_dimensions[r].height = 6

# ══════════════════════════════════════════════════════════════════════════════
# Footer note
# ══════════════════════════════════════════════════════════════════════════════
r = 28
c = ws.cell(row=r, column=1,
    value="⚑  Time inputs in Section 1 are estimates — adjust them and all downstream numbers update automatically.")
c.font = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
c.alignment = Alignment(horizontal="left", vertical="center")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.row_dimensions[r].height = 14

# Freeze panes under header
ws.freeze_panes = "A4"

# Print setup
ws.page_setup.orientation = "landscape"
ws.print_title_rows = "1:3"

path = "/Users/sohanai/.openclaw/workspace/agni-health/planning/unit-economics-panel.xlsx"
wb.save(path)
print(f"Saved: {path}")
