#!/usr/bin/env python3
"""
Agni Health Financial Model - 500 Patients in 18 Months
Corrected CAC assumptions, Hours/Week tracking, Solo physician (no MA)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Constants
TARGET_PATIENTS = 500
MONTHS = 18
MONTHLY_PRICE = 199
ANNUAL_PRICE_MONTHLY = 174  # $2,088/yr = $174/mo
BLENDED_MRR = 191.50
PAYMENT_PROCESSING = 0.029
ANNUAL_CHURN = 0.08
MONTHLY_CHURN = ANNUAL_CHURN / 12  # 0.67%
INITIAL_CONSULT = 275
PHYSICIAN_SALARY_TARGET = 300000

# Operating costs (annual)
OPERATING_COSTS = {
    'Malpractice': 10000,
    'EHR': 2000,
    'Licensing': 8000,
    'Accounting': 8000,
    'Software': 5000,
    'Insurance/Misc': 5000
}
TOTAL_ANNUAL_OPEX = sum(OPERATING_COSTS.values())  # $38,000
MONTHLY_OPEX = TOTAL_ANNUAL_OPEX / 12

# CAC assumptions (corrected)
# Months 1-6: $150, Months 7-12: $120, Months 13-18: $100
# After month 6: 20% of new patients from referrals at $0 CAC
def get_cac(month):
    if month <= 6:
        return 150
    elif month <= 12:
        return 120
    else:
        return 100

def get_referral_rate(month):
    """After month 6, 20% of new patients come from referrals at $0 CAC"""
    if month <= 6:
        return 0.0
    else:
        return 0.20

# Hours per patient per year
HOURS_PER_PATIENT_YEAR = 2.75
WEEKS_PER_YEAR = 52

# Patient growth - S-curve ramp to 500 in 18 months
# Start slow, accelerate, then maintain
MONTHLY_NEW_PATIENTS = [
    15, 18, 22, 26, 29, 32,    # Months 1-6: Building (142 total)
    34, 36, 36, 36, 34, 32,    # Months 7-12: Peak growth (208 total)
    30, 30, 29, 28, 27, 26     # Months 13-18: Steady (170 total) -> 500 after churn
]

# Adjust to hit exactly 500
total_planned = sum(MONTHLY_NEW_PATIENTS)
# We'll account for churn in the model

def create_workbook():
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    money_format = '"$"#,##0'
    money_format_decimal = '"$"#,##0.00'
    percent_format = '0.0%'
    number_format = '#,##0'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    milestone_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # ========== TAB 1: Monthly Model ==========
    ws1 = wb.active
    ws1.title = "Monthly Model"
    
    # Headers
    headers = [
        "Month", "New Patients", "Referral Patients", "Paid Patients", 
        "Churned", "Cumulative Patients", "Hours/Week Required",
        "Gross MRR", "Processing Fees", "Net MRR",
        "Initial Consults", "Total Revenue",
        "CAC Rate", "Paid CAC", "Referral CAC", "Total CAC",
        "Operating Costs", "Total Costs",
        "Net Income", "Cumulative Net"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    cumulative_patients = 0
    cumulative_net = 0
    
    for month in range(1, MONTHS + 1):
        row = month + 1
        
        # Patient calculations
        new_patients = MONTHLY_NEW_PATIENTS[month - 1]
        referral_rate = get_referral_rate(month)
        referral_patients = int(new_patients * referral_rate)
        paid_patients = new_patients - referral_patients
        
        # Churn (applied to previous month's cumulative)
        churned = int(cumulative_patients * MONTHLY_CHURN) if cumulative_patients > 0 else 0
        
        # Update cumulative
        cumulative_patients = cumulative_patients + new_patients - churned
        
        # Hours per week required
        hours_per_week = (cumulative_patients * HOURS_PER_PATIENT_YEAR) / WEEKS_PER_YEAR
        
        # Revenue
        gross_mrr = cumulative_patients * BLENDED_MRR
        processing_fees = gross_mrr * PAYMENT_PROCESSING
        net_mrr = gross_mrr - processing_fees
        initial_consults = new_patients * INITIAL_CONSULT
        total_revenue = net_mrr + initial_consults
        
        # CAC
        cac_rate = get_cac(month)
        paid_cac = paid_patients * cac_rate
        referral_cac = 0  # Referrals are free
        total_cac = paid_cac + referral_cac
        
        # Operating costs (monthly)
        operating_costs = MONTHLY_OPEX
        total_costs = total_cac + operating_costs
        
        # Net income
        net_income = total_revenue - total_costs
        cumulative_net += net_income
        
        # Write row
        data = [
            month, new_patients, referral_patients, paid_patients,
            churned, cumulative_patients, hours_per_week,
            gross_mrr, processing_fees, net_mrr,
            initial_consults, total_revenue,
            cac_rate, paid_cac, referral_cac, total_cac,
            operating_costs, total_costs,
            net_income, cumulative_net
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Number formatting
            if col == 1:  # Month
                cell.number_format = number_format
            elif col in [2, 3, 4, 5, 6]:  # Patient counts
                cell.number_format = number_format
            elif col == 7:  # Hours/week
                cell.number_format = '0.0'
            elif col in [8, 9, 10, 11, 12, 14, 15, 16, 17, 18, 19, 20]:  # Money
                cell.number_format = money_format
            elif col == 13:  # CAC rate
                cell.number_format = money_format
        
        # Highlight milestones
        if cumulative_patients >= 100 and cumulative_patients < 110:
            for c in range(1, len(headers) + 1):
                ws1.cell(row=row, column=c).fill = milestone_fill
        elif cumulative_patients >= 250 and cumulative_patients < 260:
            for c in range(1, len(headers) + 1):
                ws1.cell(row=row, column=c).fill = milestone_fill
        elif cumulative_patients >= 500:
            for c in range(1, len(headers) + 1):
                ws1.cell(row=row, column=c).fill = milestone_fill
        
        # Highlight positive/negative net income
        if net_income >= 0:
            ws1.cell(row=row, column=19).fill = positive_fill
        else:
            ws1.cell(row=row, column=19).fill = negative_fill
    
    # Totals row
    totals_row = MONTHS + 2
    ws1.cell(row=totals_row, column=1, value="TOTAL").font = header_font
    
    # Sum formulas
    sum_cols = {
        2: f"=SUM(B2:B{MONTHS+1})",   # New patients
        3: f"=SUM(C2:C{MONTHS+1})",   # Referral patients
        4: f"=SUM(D2:D{MONTHS+1})",   # Paid patients
        5: f"=SUM(E2:E{MONTHS+1})",   # Churned
        8: f"=SUM(H2:H{MONTHS+1})",   # Gross MRR
        9: f"=SUM(I2:I{MONTHS+1})",   # Processing fees
        10: f"=SUM(J2:J{MONTHS+1})",  # Net MRR
        11: f"=SUM(K2:K{MONTHS+1})",  # Initial consults
        12: f"=SUM(L2:L{MONTHS+1})",  # Total revenue
        14: f"=SUM(N2:N{MONTHS+1})",  # Paid CAC
        15: f"=SUM(O2:O{MONTHS+1})",  # Referral CAC
        16: f"=SUM(P2:P{MONTHS+1})",  # Total CAC
        17: f"=SUM(Q2:Q{MONTHS+1})",  # Operating costs
        18: f"=SUM(R2:R{MONTHS+1})",  # Total costs
        19: f"=SUM(S2:S{MONTHS+1})",  # Net income
    }
    
    for col, formula in sum_cols.items():
        cell = ws1.cell(row=totals_row, column=col, value=formula)
        cell.font = header_font
        cell.border = thin_border
        cell.number_format = money_format if col >= 8 else number_format
    
    # Column widths
    col_widths = [8, 12, 14, 12, 10, 16, 16, 12, 14, 12, 14, 14, 10, 12, 12, 12, 14, 12, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    
    # ========== TAB 2: Summary ==========
    ws2 = wb.create_sheet("Summary")
    
    # Calculate final values for summary
    final_patients = cumulative_patients
    final_mrr = final_patients * BLENDED_MRR
    final_arr = final_mrr * 12
    final_hours_week = (final_patients * HOURS_PER_PATIENT_YEAR) / WEEKS_PER_YEAR
    
    # Total revenue over 18 months
    total_revenue_18mo = 0
    total_cac_18mo = 0
    total_opex_18mo = MONTHLY_OPEX * 18
    
    cum_pts = 0
    for month in range(1, MONTHS + 1):
        new_pts = MONTHLY_NEW_PATIENTS[month - 1]
        ref_rate = get_referral_rate(month)
        ref_pts = int(new_pts * ref_rate)
        paid_pts = new_pts - ref_pts
        churned = int(cum_pts * MONTHLY_CHURN) if cum_pts > 0 else 0
        cum_pts = cum_pts + new_pts - churned
        
        gross_mrr = cum_pts * BLENDED_MRR
        proc_fees = gross_mrr * PAYMENT_PROCESSING
        net_mrr = gross_mrr - proc_fees
        consults = new_pts * INITIAL_CONSULT
        total_revenue_18mo += net_mrr + consults
        
        cac_rate = get_cac(month)
        total_cac_18mo += paid_pts * cac_rate
    
    total_costs_18mo = total_cac_18mo + total_opex_18mo
    net_income_18mo = total_revenue_18mo - total_costs_18mo
    
    summary_data = [
        ("AGNI HEALTH - 18-MONTH FINANCIAL SUMMARY", ""),
        ("", ""),
        ("TARGET METRICS", ""),
        ("Target Patients", 500),
        ("Final Patient Count", final_patients),
        ("Achievement %", final_patients / 500),
        ("", ""),
        ("PHYSICIAN WORKLOAD", ""),
        ("Hours per Patient per Year", HOURS_PER_PATIENT_YEAR),
        ("Final Hours/Week Required", final_hours_week),
        ("", ""),
        ("REVENUE (Month 18)", ""),
        ("Monthly MRR", final_mrr),
        ("Annualized Revenue", final_arr),
        ("", ""),
        ("18-MONTH TOTALS", ""),
        ("Total Revenue", total_revenue_18mo),
        ("Total CAC Spend", total_cac_18mo),
        ("Total Operating Costs", total_opex_18mo),
        ("Total Net Income", net_income_18mo),
        ("", ""),
        ("CAC ASSUMPTIONS", ""),
        ("Months 1-6 CAC", "$150/patient"),
        ("Months 7-12 CAC", "$120/patient"),
        ("Months 13-18 CAC", "$100/patient"),
        ("Referral Rate (after M6)", "20% at $0 CAC"),
        ("", ""),
        ("UNIT ECONOMICS", ""),
        ("Blended MRR/Patient", BLENDED_MRR),
        ("Initial Consultation", INITIAL_CONSULT),
        ("Annual Churn Rate", ANNUAL_CHURN),
        ("Payment Processing", PAYMENT_PROCESSING),
        ("", ""),
        ("OPERATING COSTS (Annual)", ""),
        ("Malpractice", OPERATING_COSTS['Malpractice']),
        ("EHR", OPERATING_COSTS['EHR']),
        ("Licensing", OPERATING_COSTS['Licensing']),
        ("Accounting", OPERATING_COSTS['Accounting']),
        ("Software", OPERATING_COSTS['Software']),
        ("Insurance/Misc", OPERATING_COSTS['Insurance/Misc']),
        ("Total Annual OpEx", TOTAL_ANNUAL_OPEX),
        ("", ""),
        ("PHYSICIAN COMPENSATION TARGET", ""),
        ("Annual Salary Target", PHYSICIAN_SALARY_TARGET),
        ("Monthly Salary Target", PHYSICIAN_SALARY_TARGET / 12),
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        
        if label in ["AGNI HEALTH - 18-MONTH FINANCIAL SUMMARY", "TARGET METRICS", 
                     "PHYSICIAN WORKLOAD", "REVENUE (Month 18)", "18-MONTH TOTALS",
                     "CAC ASSUMPTIONS", "UNIT ECONOMICS", "OPERATING COSTS (Annual)",
                     "PHYSICIAN COMPENSATION TARGET"]:
            cell_a.font = Font(bold=True, size=12)
            if label == "AGNI HEALTH - 18-MONTH FINANCIAL SUMMARY":
                cell_a.font = Font(bold=True, size=14)
        
        # Format values
        if isinstance(value, (int, float)) and value != "":
            if label in ["Achievement %", "Annual Churn Rate", "Payment Processing"]:
                cell_b.number_format = percent_format
            elif "Hours" in label:
                cell_b.number_format = '0.0'
            elif value >= 1000:
                cell_b.number_format = money_format
            elif "MRR" in label or "Consultation" in label:
                cell_b.number_format = money_format_decimal
    
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 20
    
    # ========== TAB 3: Marketing Ramp ==========
    ws3 = wb.create_sheet("Marketing Ramp")
    
    # Headers
    marketing_headers = [
        "Month", "Phase", "New Patients", "Referral %", "Paid Patients",
        "CAC Rate", "CAC Spend", "Cumulative CAC", "Avg CAC/Patient"
    ]
    
    for col, header in enumerate(marketing_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    cumulative_cac = 0
    cumulative_patients_mkt = 0
    
    for month in range(1, MONTHS + 1):
        row = month + 1
        
        new_patients = MONTHLY_NEW_PATIENTS[month - 1]
        referral_rate = get_referral_rate(month)
        referral_patients = int(new_patients * referral_rate)
        paid_patients = new_patients - referral_patients
        cac_rate = get_cac(month)
        cac_spend = paid_patients * cac_rate
        cumulative_cac += cac_spend
        cumulative_patients_mkt += new_patients
        avg_cac = cumulative_cac / cumulative_patients_mkt
        
        # Determine phase
        if month <= 6:
            phase = "Launch & Build"
        elif month <= 12:
            phase = "Growth & Referrals"
        else:
            phase = "Optimization"
        
        data = [
            month, phase, new_patients, referral_rate, paid_patients,
            cac_rate, cac_spend, cumulative_cac, avg_cac
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws3.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            if col == 4:  # Referral %
                cell.number_format = percent_format
            elif col in [6, 7, 8]:  # CAC amounts
                cell.number_format = money_format
            elif col == 9:  # Avg CAC
                cell.number_format = money_format_decimal
    
    # Totals row
    totals_row = MONTHS + 2
    ws3.cell(row=totals_row, column=1, value="TOTAL").font = header_font
    ws3.cell(row=totals_row, column=3, value=f"=SUM(C2:C{MONTHS+1})").number_format = number_format
    ws3.cell(row=totals_row, column=5, value=f"=SUM(E2:E{MONTHS+1})").number_format = number_format
    ws3.cell(row=totals_row, column=7, value=f"=SUM(G2:G{MONTHS+1})").number_format = money_format
    
    for col in [1, 3, 5, 7]:
        ws3.cell(row=totals_row, column=col).font = header_font
        ws3.cell(row=totals_row, column=col).border = thin_border
    
    # Column widths
    mkt_widths = [8, 18, 14, 12, 14, 10, 12, 14, 14]
    for i, width in enumerate(mkt_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = width
    
    # Save workbook
    output_path = "/Users/sohanai/.openclaw/workspace/agni-health-500-18mo.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    print(f"\nKey Metrics:")
    print(f"  Final patients: {final_patients}")
    print(f"  Final Hours/Week: {final_hours_week:.1f}")
    print(f"  Final MRR: ${final_mrr:,.0f}")
    print(f"  Final ARR: ${final_arr:,.0f}")
    print(f"  18-Month Net Income: ${net_income_18mo:,.0f}")
    print(f"  Total CAC Spend: ${total_cac_18mo:,.0f}")
    print(f"  Avg CAC/Patient: ${total_cac_18mo/sum(MONTHLY_NEW_PATIENTS):.2f}")

if __name__ == "__main__":
    create_workbook()
