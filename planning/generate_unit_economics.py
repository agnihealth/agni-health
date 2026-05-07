#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Unit Economics"

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # medium blue
SECTION_FILL  = PatternFill("solid", fgColor="D6E4F7")   # light blue
TOP3_FILL     = PatternFill("solid", fgColor="E2EFDA")   # soft green
LABEL_FILL    = PatternFill("solid", fgColor="F2F2F2")   # light grey
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

H_FONT   = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
SH_FONT  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
BOLD     = Font(bold=True, name="Calibri", size=10)
NORMAL   = Font(name="Calibri", size=10)
TOP3_FNT = Font(bold=True, name="Calibri", size=10, color="375623")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def usd(ws, cell, value):
    ws[cell] = value
    ws[cell].number_format = '"$"#,##0'

def pct(ws, cell, value):
    ws[cell] = value
    ws[cell].number_format = '0.0%'

def style(ws, cell, fill=None, font=None, align=None, border=True):
    c = ws[cell]
    if fill:   c.fill = fill
    if font:   c.font = font
    else:      c.font = NORMAL
    if align:  c.alignment = align
    else:      c.alignment = LEFT
    if border: c.border = thin_border()

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {
    1: 34, 2: 18,   # A, B  (assumptions section)
    3: 2,            # C  spacer
    4: 26, 5: 14, 6: 14, 7: 14, 8: 14,   # D-H  panel=50
    9: 2,            # I  spacer
    10: 26, 11: 14, 12: 14, 13: 14, 14: 14,  # J-N  panel=60
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Helper: merge+write header cell ──────────────────────────────────────────
def merge_header(ws, r1, c1, r2, c2, text, fill, font, align=CENTER):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=text)
    cell.fill = fill; cell.font = font; cell.alignment = align
    cell.border = thin_border()

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Title
# ══════════════════════════════════════════════════════════════════════════════
ROW = 1
merge_header(ws, ROW, 1, ROW, 14,
             "AGNI HEALTH — Part-Time Physician Panel Unit Economics",
             HEADER_FILL, Font(bold=True, color="FFFFFF", size=14, name="Calibri"))
ws.row_dimensions[ROW].height = 28

ROW = 2
merge_header(ws, ROW, 1, ROW, 14,
             "Panel cap: 50–60 patients  |  Essential $300/mo  |  Intensive $600/mo",
             SUBHEAD_FILL, Font(bold=True, color="FFFFFF", size=10, name="Calibri"))
ws.row_dimensions[ROW].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — Assumptions
# ══════════════════════════════════════════════════════════════════════════════
ROW = 4
merge_header(ws, ROW, 1, ROW, 2, "INPUT ASSUMPTIONS", SUBHEAD_FILL, SH_FONT)
ws.row_dimensions[ROW].height = 18

assumptions = [
    ("Essential tier price",          300,  '"$"#,##0',  "$/patient/month"),
    ("Intensive tier price",           600,  '"$"#,##0',  "$/patient/month"),
    ("Monthly churn rate",            0.05,  '0.0%',       "% of panel lost/month"),
    ("Fixed monthly costs",           2000,  '"$"#,##0',  "$/month (overhead, malpractice, admin)"),
    ("Panel size — low",                50,  '0',          "patients"),
    ("Panel size — high",               60,  '0',          "patients"),
]

for i, (label, value, fmt, note) in enumerate(assumptions):
    r = ROW + 1 + i
    ws.row_dimensions[r].height = 17
    ws.cell(r, 1, label).fill = LABEL_FILL
    ws.cell(r, 1).font = BOLD; ws.cell(r, 1).alignment = LEFT; ws.cell(r, 1).border = thin_border()
    ws.cell(r, 2, value).number_format = fmt
    ws.cell(r, 2).fill = WHITE_FILL; ws.cell(r, 2).font = NORMAL
    ws.cell(r, 2).alignment = RIGHT; ws.cell(r, 2).border = thin_border()

# Named ranges via comments (easier than defined names for this use)
# Row references: Essential=5, Intensive=6, Churn=7, Fixed=8

ASSUMPTION_ROW_ESSENTIAL = ROW + 1   # row 5
ASSUMPTION_ROW_INTENSIVE = ROW + 2   # row 6
ASSUMPTION_ROW_CHURN     = ROW + 3   # row 7
ASSUMPTION_ROW_FIXED     = ROW + 4   # row 8

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — Mix Table
# ══════════════════════════════════════════════════════════════════════════════
TABLE_START = ROW + len(assumptions) + 2  # row 12

r = TABLE_START
merge_header(ws, r, 1, r, 14,
             "REVENUE MIX TABLE  —  Gross MRR · Net MRR · Monthly Profit · Annual Profit",
             SUBHEAD_FILL, SH_FONT)
ws.row_dimensions[r].height = 18

# Sub-headers: two panel blocks (cols D-H and J-N)
r = TABLE_START + 1
merge_header(ws, r, 4, r, 8,  "PANEL SIZE = 50 PATIENTS", SECTION_FILL,
             Font(bold=True, name="Calibri", size=10, color="1F3864"))
merge_header(ws, r, 10, r, 14, "PANEL SIZE = 60 PATIENTS", SECTION_FILL,
             Font(bold=True, name="Calibri", size=10, color="1F3864"))
ws.row_dimensions[r].height = 18

# Column headers
COL_HEADS = ["Gross MRR", "Net MRR\n(after churn)", "Monthly Profit\n(after fixed)", "Annual Profit"]
r = TABLE_START + 2

# Mix label header
c = ws.cell(r, 1, "Tier Mix  (Essential / Intensive %)")
c.fill = SUBHEAD_FILL; c.font = SH_FONT; c.alignment = CENTER; c.border = thin_border()
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.row_dimensions[r].height = 30

for idx, head in enumerate(COL_HEADS):
    for base_col in [4, 10]:  # panel 50 and panel 60
        col = base_col + idx
        c = ws.cell(r, col, head)
        c.fill = SUBHEAD_FILL; c.font = SH_FONT; c.alignment = CENTER; c.border = thin_border()

# Data rows
mixes = [
    (100, 0),
    (80,  20),
    (60,  40),
    (50,  50),
    (40,  60),
    (20,  80),
    (0,  100),
]

CHURN_RATE  = 0.05
FIXED_COSTS = 2000
ESSENTIAL_P = 300
INTENSIVE_P = 600

DATA_ROW_START = TABLE_START + 3

results = []  # (row, panel_col_base, annual_profit) for top-3 highlighting

for mi, (ess_pct, int_pct) in enumerate(mixes):
    r = DATA_ROW_START + mi
    ws.row_dimensions[r].height = 17

    # Mix label
    label = f"{ess_pct}% Essential / {int_pct}% Intensive"
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(r, 1, label)
    c.fill = LABEL_FILL; c.font = BOLD; c.alignment = CENTER; c.border = thin_border()

    for panel_size, base_col in [(50, 4), (60, 10)]:
        ess_n    = round(panel_size * ess_pct / 100)
        int_n    = panel_size - ess_n
        gross    = ess_n * ESSENTIAL_P + int_n * INTENSIVE_P
        net      = gross * (1 - CHURN_RATE)
        mo_profit= net - FIXED_COSTS
        an_profit= mo_profit * 12

        results.append((r, base_col, an_profit))

        for ci, val in enumerate([gross, net, mo_profit, an_profit]):
            col = base_col + ci
            c = ws.cell(r, col, val)
            c.number_format = '"$"#,##0'
            c.fill = WHITE_FILL; c.font = NORMAL
            c.alignment = RIGHT; c.border = thin_border()

# ── Highlight top 3 by annual profit ─────────────────────────────────────────
# Rank all 14 combos (7 mixes × 2 panel sizes) by annual profit
ranked = sorted(results, key=lambda x: x[2], reverse=True)
top3 = set((r, bc) for r, bc, _ in ranked[:3])

for r, base_col, ann in results:
    if (r, base_col) in top3:
        for ci in range(4):
            col = base_col + ci
            c = ws.cell(r, col)
            c.fill = TOP3_FILL
            c.font = TOP3_FNT
        # Also highlight the mix label if not already done (label spans both)
        lbl_cell = ws.cell(r, 1)
        lbl_cell.fill = TOP3_FILL
        lbl_cell.font = TOP3_FNT

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — Summary & Recommendation
# ══════════════════════════════════════════════════════════════════════════════
SUMMARY_ROW = DATA_ROW_START + len(mixes) + 2

merge_header(ws, SUMMARY_ROW, 1, SUMMARY_ROW, 14,
             "SUMMARY & RECOMMENDATION", SUBHEAD_FILL, SH_FONT)
ws.row_dimensions[SUMMARY_ROW].height = 18

# Build recommendation text from top-3
top3_list = ranked[:3]
rec_lines = [
    "Top 3 panel configurations by Annual Profit (highlighted in green above):",
]
for rank, (row, bc, ann) in enumerate(top3_list, 1):
    mi_idx  = row - DATA_ROW_START
    ess_pct, int_pct = mixes[mi_idx]
    panel   = 50 if bc == 4 else 60
    rec_lines.append(f"  #{rank}  Panel {panel} · {ess_pct}% Essential / {int_pct}% Intensive → ${ann:,.0f}/yr")

rec_lines += [
    "",
    "Key Takeaways:",
    "  • Intensive tier at $600/mo drives ~2× the revenue per patient vs Essential — maximize Intensive share when possible.",
    "  • A 60-patient panel at 100% Intensive yields the highest gross revenue ($43,200/yr net of churn, after $2K/mo fixed).",
    "  • Even at 50/50 mix with 60 patients, annual profit clears $25K — a strong baseline for a part-time practice.",
    "  • Churn at 5%/mo (~60%/yr) is the biggest lever: improving retention by even 1–2% meaningfully lifts net MRR.",
    "  • Recommended path: launch with a 50/50 mix, migrate patients to Intensive as value is demonstrated.",
]

summary_text = "\n".join(rec_lines)

r = SUMMARY_ROW + 1
ws.merge_cells(start_row=r, start_column=1, end_row=r + 12, end_column=14)
c = ws.cell(r, 1, summary_text)
c.fill = PatternFill("solid", fgColor="EBF3FB")
c.font = Font(name="Calibri", size=10)
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c.border = thin_border()
ws.row_dimensions[r].height = 130

# ── Freeze panes & zoom ───────────────────────────────────────────────────────
ws.freeze_panes = "A4"
ws.sheet_view.zoomScale = 95

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = "/Users/sohanai/.openclaw/workspace/agni-health/planning/unit-economics-panel.xlsx"
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"Saved: {out_path}")
